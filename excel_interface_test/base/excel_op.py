# --------------------------------------------------
# !/usr/bin/python
# -*- coding: utf-8 -*-
# PN: DBInterface
# FN: excel_op
# Author: xiaxu
# DATA: 2021/11/23
# Description:操作excel的基本模块
# ---------------------------------------------------
import openpyxl,re
from openpyxl.styles import Font,PatternFill
from excel_interface_test.base.function import  project_path


class Singleton:
    #单例模式
    def __new__(cls, *args, **kwargs):
        if not hasattr(cls,'_instance'):
            cls._instance = super(Singleton,cls).__new__(cls)
        return cls._instance

class ExcelOP(Singleton):
    def __init__(self,sheet_name):
        """
        初始化读取excel文件；读取指定的数据表
        :param sheet_name: 数据表名
        """
        self.xl_path = project_path()+"\data\\api_cases.xlsx"
        print(self.xl_path)
        self.xl= openpyxl.load_workbook(self.xl_path,read_only=False)
        self.sheet = self.xl[sheet_name]

    def open_sheet(self,name):
        self.sheet = self.xl[name]

    def get_rows(self):
        #按照行返回所有数据
        rows = []
        for row in self.sheet.iter_rows():  # 按照行迭代
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            rows.append(tuple(row_list))
        return rows

    def get_cols(self):
        #按照列返回所有数据
        cols = []
        for col in self.sheet.iter_cols():  # 按照列迭代
            col_list = []
            for cell in col:  # 打印一列的每个单元格内容
                col_list.append(cell.value)
            cols.append(tuple(col_list))
        return cols

    def get_record(self,record_str):
        #将record解析称为列表
        parttern = '.+\;{1}' #非贪婪模式,改为了贪婪模式，性能更好
        return record_str.split(';')
    def write_result(self,row,columm,status=1):
        if status ==1:
            self.sheet.cell(row,columm,value = "Pass")
            self.sheet.cell(row, columm).fill = PatternFill("solid",fgColor="AACF91")
            self.sheet.cell(row, columm).font = Font(bold=True)
        elif status == 2:
            self.sheet.cell(row, columm,"Fail")
            self.sheet.cell(row, columm).fill = PatternFill("solid", fgColor="FF0000")
            self.sheet.cell(row, columm).font = Font(bold=True)
        elif status ==3:
            self.sheet.cell(row, columm).value = "Blcoked"
            self.sheet.cell(row, columm).fill = PatternFill("solid", fgColor="CDCDCD")#浅灰色
            self.sheet.cell(row, columm).font = Font(bold=True)
        elif status ==4:
            #初始化单元格
            self.sheet.cell(row, columm).value = ""
            self.sheet.cell(row, columm).fill = PatternFill("solid", fgColor="000000")#浅灰色
            self.sheet.cell(row, columm).font = Font(bold=False)
        elif status ==5:
            #测试发生错误
            self.sheet.cell(row, columm).value = "Error"
            self.sheet.cell(row, columm).fill = PatternFill("solid", fgColor="00FF00")#浅灰色
            self.sheet.cell(row, columm).font = Font(bold=True)
        self.xl.save(self.xl_path)

    def write_data(self,row,col,data):
        """定义写入数据的方法"""
        self.sheet.cell(row, col, value=data)
        self.xl.save(self.xl_path)



if __name__ == '__main__':
    xl = ExcelOP("test_cases")
    params_sheet = xl.open_sheet('params') #重写数据表
    params_con = eval(xl.get_rows()[1][0])#获取参数的内容返回一个可迭代对象
    print(params_con['read'],type(params_con['read']))
    for keyword in list(params_con['read']):
        print(keyword,type(keyword))





